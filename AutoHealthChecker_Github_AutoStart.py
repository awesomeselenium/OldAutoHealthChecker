from selenium import webdriver
from openpyxl import load_workbook
import time
startnumber_manual=1
driver = webdriver.Chrome('chromedriver.exe')
load_wb = load_workbook(r'studentlist.xlsx',data_only=True)
load_ws = load_wb['Sheet1']
lastnumber = 1
autostart = 1
while (load_ws.cell(lastnumber,1).value != None):
    lastnumber=lastnumber+1
while (load_ws.cell(autostart,3).value != None):
    autostart=autostart+1
print("StartPoint:", autostart)
startnumber_manual= autostart

print("엑셀 파일의 학생은 자신의 건강상태와 주위 사람들의 건강 상태를 점검하여서 문제가 없었음을 인정하였습니다. ")

for cnt in range(startnumber_manual,lastnumber):
    driver.get("https://eduro.pen.go.kr/stv_cvd_co00_002.do")  
    driver.implicitly_wait(10)
    driver.find_element_by_xpath('//*[@id="schulNm"]').send_keys('여기에 학교이름 입력')
    driver.find_element_by_xpath('//*[@id="btnSrchSchul"]').click()
    time.sleep(2)
    driver.find_element_by_xpath('//*[@id="pName"]').send_keys(load_ws.cell(cnt,1).value)
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="frnoRidno"]').send_keys(load_ws.cell(cnt,2).value)
    driver.implicitly_wait(10)
    time.sleep(1)
    btc = driver.find_element_by_xpath('//*[@id="btnConfirm"]')
    driver.find_element_by_xpath('//*[@id="btnConfirm"]').click()
    driver.implicitly_wait(10)
    time.sleep(2)
    driver.find_element_by_xpath('//*[@id="rspns011"]').click()
    driver.find_element_by_xpath('//*[@id="rspns02"]').click()
    driver.find_element_by_xpath('//*[@id="rspns070"]').click()
    driver.find_element_by_xpath('//*[@id="rspns080"]').click()
    driver.find_element_by_xpath('//*[@id="rspns090"]').click()
    time.sleep(2)
    driver.find_element_by_xpath('//*[@id="btnConfirm"]').click()
    print(load_ws.cell(cnt,1).value)
    print(cnt,"/",lastnumber-1,"completed")
    load_ws.cell(row=cnt,column=3).value = 1
    load_wb.save(r'studentlist.xlsx')

for cnt in range(1,lastnumber) :
    load_ws.cell(row=cnt,column=3).value = ""

load_wb.save(r'studentlist.xlsx')
    
